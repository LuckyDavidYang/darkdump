import importlib
import io
import json
import os
import sys
import tempfile
import unittest
import warnings
from unittest.mock import patch


class FakeResponse:
    def __init__(self, status_code=200, content="", json_data=None):
        self.status_code = status_code
        self.content = content.encode("utf-8") if isinstance(content, str) else content
        self._json_data = json_data or {}

    def json(self):
        return self._json_data


HOMEPAGE_HTML = """
<html>
  <body>
    <form id="searchForm">
      <input type="hidden" name="nonce" value="abc123">
    </form>
  </body>
</html>
"""

SEARCH_RESULTS_HTML = """
<div id="ahmiaResultsPage">
  <li class="result">
    <a>Alpha Market</a>
    <cite>alpha.onion</cite>
    <p>Primary listing</p>
  </li>
  <li class="result">
    <a>Beta Market</a>
    <cite>http://beta.onion</cite>
    <p>Secondary listing</p>
  </li>
</div>
"""

SEARCH_RESULTS_HTML_THREE = """
<div id="ahmiaResultsPage">
  <li class="result"><a>One</a><cite>one.onion</cite><p>One desc</p></li>
  <li class="result"><a>Two</a><cite>two.onion</cite><p>Two desc</p></li>
  <li class="result"><a>Three</a><cite>three.onion</cite><p>Three desc</p></li>
</div>
"""

SITE_HTML = """
<html>
  <head>
    <meta name="description" content="test metadata">
  </head>
  <body>
    <a href="https://example.com">Example</a>
    <a href="docs/report.pdf">Report</a>
    Contact: ops@alpha.onion
  </body>
</html>
"""

SITE_HTML_BETA = """
<html>
  <head>
    <meta property="og:title" content="beta metadata">
  </head>
  <body>
    <a href="https://beta.example.com">Beta</a>
    Reach us at admin@beta.onion
  </body>
</html>
"""


def get_collect_dark_net():
    module = importlib.import_module("darkdump_collector")
    return module.collect_dark_net


def get_darkdump_collector_module():
    return importlib.import_module("darkdump_collector")


class FakeFuture:
    def __init__(self, result=None, error=None):
        self._result = result
        self._error = error

    def result(self):
        if self._error is not None:
            raise self._error
        return self._result


class FakeProcessPoolExecutor:
    def __init__(self, max_workers, future_factory):
        self.max_workers = max_workers
        self.future_factory = future_factory
        self.submitted = []

    def submit(self, fn, *args, **kwargs):
        future = self.future_factory(fn, *args, **kwargs)
        self.submitted.append((fn, args, kwargs, future))
        return future

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc, tb):
        return False


class CollectDarkNetTests(unittest.TestCase):
    def test_importing_darkdump_collector_has_no_syntax_warnings(self):
        for module_name in ("darkdump_collector", "darkdump"):
            sys.modules.pop(module_name, None)

        with warnings.catch_warnings(record=True) as caught_warnings:
            warnings.simplefilter("always")
            collect_dark_net = get_collect_dark_net()

        self.assertTrue(callable(collect_dark_net))
        syntax_warnings = [warning for warning in caught_warnings if issubclass(warning.category, SyntaxWarning)]
        self.assertEqual(syntax_warnings, [])

    @patch("darkdump.Darkdump.analyze_text")
    @patch("darkdump.Darkdump.extract_keywords")
    @patch("darkdump.Platform.get_tor_connection_status", return_value=(True, "185.220.101.1"))
    @patch("darkdump.requests.get")
    def test_collect_dark_net_returns_structured_results(
        self,
        mock_get,
        _mock_tor_status,
        mock_extract_keywords,
        mock_analyze_text,
    ):
        mock_extract_keywords.side_effect = [["alpha", "market"], ["beta", "market"]]
        mock_analyze_text.side_effect = [
            {"top_words": [("alpha", 2)], "sentiment": {"polarity": 0.4, "subjectivity": 0.2}},
            {"top_words": [("beta", 2)], "sentiment": {"polarity": 0.1, "subjectivity": 0.3}},
        ]
        mock_get.side_effect = [
            FakeResponse(content=HOMEPAGE_HTML),
            FakeResponse(content=SEARCH_RESULTS_HTML),
            FakeResponse(content=SITE_HTML),
            FakeResponse(content=SITE_HTML_BETA),
        ]

        collect_dark_net = get_collect_dark_net()
        with patch("sys.stdout", new_callable=io.StringIO) as fake_stdout:
            result = collect_dark_net(" markets ", 2)

        self.assertEqual(result["query"], "markets")
        self.assertEqual(result["requested_amount"], 2)
        self.assertEqual(result["returned_count"], 2)
        self.assertTrue(result["proxy_enabled"])
        self.assertTrue(result["scrape_enabled"])
        self.assertFalse(result["images_enabled"])
        self.assertTrue(result["tor_checked"])
        self.assertTrue(result["tor_ok"])
        self.assertEqual(result["tor_ip"], "Current IP Address via Tor: 185.220.101.1")
        self.assertEqual(result["errors"], [])
        self.assertEqual(len(result["results"]), 2)

        first_result = result["results"][0]
        self.assertEqual(first_result["index"], 1)
        self.assertEqual(first_result["title"], "Alpha Market")
        self.assertEqual(first_result["description"], "Primary listing")
        self.assertEqual(first_result["onion_link"], "http://alpha.onion")
        self.assertEqual(first_result["keywords"], ["alpha", "market"])
        self.assertEqual(first_result["sentiment"], {"polarity": 0.4, "subjectivity": 0.2})
        self.assertEqual(first_result["metadata"], {"description": "test metadata"})
        self.assertEqual(first_result["link_count"], 2)
        self.assertIn("https://example.com", first_result["links"])
        self.assertIn("ops@alpha.onion", first_result["emails"])
        self.assertEqual(first_result["documents"], ["docs/report.pdf"])

        stdout = fake_stdout.getvalue()
        self.assertIn("query=markets", stdout)
        self.assertIn("returned_count=2", stdout)
        self.assertIn("tor_ip=Current IP Address via Tor: 185.220.101.1", stdout)
        self.assertIn("error_count=0", stdout)

    def test_collect_dark_net_rejects_invalid_keyword(self):
        collect_dark_net = get_collect_dark_net()

        with self.assertRaises(TypeError):
            collect_dark_net(None, 1)

        with self.assertRaises(ValueError):
            collect_dark_net("   ", 1)

    def test_collect_dark_net_rejects_invalid_amount(self):
        collect_dark_net = get_collect_dark_net()

        with self.assertRaises(TypeError):
            collect_dark_net("markets", "1")

        with self.assertRaises(TypeError):
            collect_dark_net("markets", True)

        with self.assertRaises(ValueError):
            collect_dark_net("markets", 0)

    def test_collect_dark_net_rejects_invalid_retry_times(self):
        collect_dark_net = get_collect_dark_net()

        with self.assertRaises(TypeError):
            collect_dark_net("markets", 1, retry_times="3")

        with self.assertRaises(TypeError):
            collect_dark_net("markets", 1, retry_times=True)

        with self.assertRaises(ValueError):
            collect_dark_net("markets", 1, retry_times=-1)

    @patch("darkdump_collector.Darkdump.collect")
    def test_collect_dark_net_retries_runtime_errors_until_success(self, mock_collect):
        collect_dark_net = get_collect_dark_net()
        expected_result = {
            "query": "markets",
            "requested_amount": 1,
            "returned_count": 0,
            "proxy_enabled": True,
            "scrape_enabled": True,
            "images_enabled": False,
            "tor_checked": True,
            "tor_ok": True,
            "tor_ip": "Current IP Address via Tor: 185.220.101.1",
            "errors": [],
            "results": [],
        }
        mock_collect.side_effect = [
            RuntimeError("temporary failure"),
            RuntimeError("temporary failure"),
            expected_result,
        ]

        with patch("sys.stdout", new_callable=io.StringIO) as fake_stdout:
            result = collect_dark_net("markets", 1)

        self.assertEqual(result, expected_result)
        self.assertEqual(mock_collect.call_count, 3)
        stdout = fake_stdout.getvalue()
        self.assertEqual(stdout.count("query=markets"), 1)
        self.assertIn("error_count=0", stdout)

    @patch("darkdump_collector.Darkdump.collect")
    def test_collect_dark_net_raises_after_exhausting_default_retries(self, mock_collect):
        collect_dark_net = get_collect_dark_net()
        mock_collect.side_effect = RuntimeError("persistent failure")

        with patch("sys.stdout", new_callable=io.StringIO) as fake_stdout:
            with self.assertRaises(RuntimeError) as context:
                collect_dark_net("markets", 1)

        self.assertIn("persistent failure", str(context.exception))
        self.assertEqual(mock_collect.call_count, 4)
        stdout = fake_stdout.getvalue()
        self.assertIn("query=markets", stdout)
        self.assertIn("returned_count=0", stdout)
        self.assertIn("tor_ip=None", stdout)
        self.assertIn("error_count=1", stdout)
        self.assertIn("persistent failure", stdout)

    @patch("darkdump.Platform.get_tor_connection_status", return_value=(False, None))
    @patch("darkdump.requests.get")
    def test_collect_dark_net_raises_when_tor_check_fails(self, mock_get, _mock_tor_status):
        mock_get.side_effect = [
            FakeResponse(content=HOMEPAGE_HTML),
            FakeResponse(content=SEARCH_RESULTS_HTML),
        ]

        collect_dark_net = get_collect_dark_net()
        with self.assertRaises(RuntimeError) as context:
            collect_dark_net("markets", 2, retry_times=0)

        self.assertIn("Tor", str(context.exception))

    @patch("darkdump.requests.get")
    def test_collect_dark_net_raises_on_ahmia_homepage_failure(self, mock_get):
        mock_get.return_value = FakeResponse(status_code=500)

        collect_dark_net = get_collect_dark_net()
        with self.assertRaises(RuntimeError) as context:
            collect_dark_net("markets", 1)

        self.assertIn("Couldn't fetch", str(context.exception))

    @patch("darkdump.requests.get")
    def test_collect_dark_net_raises_when_nonce_is_missing(self, mock_get):
        mock_get.return_value = FakeResponse(content="<html><body></body></html>")

        collect_dark_net = get_collect_dark_net()
        with self.assertRaises(RuntimeError) as context:
            collect_dark_net("markets", 1)

        self.assertIn("nonce", str(context.exception))

    @patch("darkdump.requests.get")
    def test_collect_dark_net_raises_when_results_container_is_missing(self, mock_get):
        mock_get.side_effect = [
            FakeResponse(content=HOMEPAGE_HTML),
            FakeResponse(content="<html><body>No results container</body></html>"),
        ]

        collect_dark_net = get_collect_dark_net()
        with self.assertRaises(RuntimeError) as context:
            collect_dark_net("markets", 1, retry_times=0)

        self.assertIn("extract results", str(context.exception))

    @patch("darkdump.Darkdump.analyze_text", return_value={"top_words": [], "sentiment": {"polarity": 0.0, "subjectivity": 0.0}})
    @patch("darkdump.Darkdump.extract_keywords", return_value=["survivor"])
    @patch("darkdump.Platform.get_tor_connection_status", return_value=(True, "185.220.101.2"))
    @patch("darkdump.requests.get")
    def test_collect_dark_net_records_site_errors_and_continues(
        self,
        mock_get,
        _mock_tor_status,
        _mock_extract_keywords,
        _mock_analyze_text,
    ):
        mock_get.side_effect = [
            FakeResponse(content=HOMEPAGE_HTML),
            FakeResponse(content=SEARCH_RESULTS_HTML),
            RuntimeError("dead onion"),
            FakeResponse(content=SITE_HTML_BETA),
        ]

        collect_dark_net = get_collect_dark_net()
        result = collect_dark_net("markets", 2)

        self.assertEqual(result["returned_count"], 1)
        self.assertEqual(len(result["results"]), 1)
        self.assertEqual(len(result["errors"]), 1)
        self.assertEqual(result["tor_ip"], "Current IP Address via Tor: 185.220.101.2")
        self.assertEqual(result["errors"][0]["stage"], "site")
        self.assertEqual(result["errors"][0]["url"], "http://alpha.onion")
        self.assertIn("dead onion", result["errors"][0]["message"])

    @patch("darkdump.Darkdump.analyze_text", return_value={"top_words": [], "sentiment": {"polarity": 0.0, "subjectivity": 0.0}})
    @patch("darkdump.Darkdump.extract_keywords", return_value=["limited"])
    @patch("darkdump.Platform.get_tor_connection_status", return_value=(True, "185.220.101.3"))
    @patch("darkdump.requests.get")
    def test_collect_dark_net_limits_results_to_requested_amount(
        self,
        mock_get,
        _mock_tor_status,
        _mock_extract_keywords,
        _mock_analyze_text,
    ):
        mock_get.side_effect = [
            FakeResponse(content=HOMEPAGE_HTML),
            FakeResponse(content=SEARCH_RESULTS_HTML_THREE),
            FakeResponse(content=SITE_HTML),
            FakeResponse(content=SITE_HTML_BETA),
            FakeResponse(content=SITE_HTML),
        ]

        collect_dark_net = get_collect_dark_net()
        result = collect_dark_net("markets", 2)

        self.assertEqual(result["returned_count"], 2)
        self.assertEqual(len(result["results"]), 2)
        self.assertEqual(mock_get.call_count, 4)

    @patch("darkdump.Platform.get_tor_connection_status", side_effect=RuntimeError("tor unavailable"))
    @patch("darkdump.requests.get")
    def test_collect_dark_net_raises_when_tor_status_lookup_errors(self, mock_get, _mock_tor_status):
        mock_get.side_effect = [
            FakeResponse(content=HOMEPAGE_HTML),
            FakeResponse(content=SEARCH_RESULTS_HTML),
        ]

        collect_dark_net = get_collect_dark_net()
        with self.assertRaises(RuntimeError) as context:
            collect_dark_net("markets", 2, retry_times=0)

        self.assertIn("Tor", str(context.exception))

    def test_batch_collect_dark_net_rejects_invalid_processes(self):
        module = get_darkdump_collector_module()

        with self.assertRaises(TypeError):
            module.batch_collect_dark_net(["alpha"], 1, processes="10")

        with self.assertRaises(TypeError):
            module.batch_collect_dark_net(["alpha"], 1, processes=True)

        with self.assertRaises(ValueError):
            module.batch_collect_dark_net(["alpha"], 1, processes=0)

    @patch("darkdump_collector._batch_collect_worker")
    def test_batch_collect_dark_net_returns_success_items_in_input_order_and_prints_completion_order(self, mock_worker):
        module = get_darkdump_collector_module()
        executors = []

        def worker_side_effect(index, key_word, amount, retry_times):
            return {
                "index": index,
                "item": {
                    "collected_date": f"2026-04-08",
                    "collected_time": f"10:00:0{index}",
                    "search_keyword": key_word,
                    "status": "success",
                    "collect_result": {
                        "query": key_word,
                        "requested_amount": amount,
                        "returned_count": 1,
                        "proxy_enabled": True,
                        "scrape_enabled": True,
                        "images_enabled": False,
                        "tor_checked": True,
                        "tor_ok": True,
                        "tor_ip": f"Current IP Address via Tor: 185.220.101.{index + 1}",
                        "errors": [],
                        "results": [{"index": 1, "title": key_word.title()}],
                    },
                },
            }

        mock_worker.side_effect = worker_side_effect

        def future_factory(fn, *args, **kwargs):
            return FakeFuture(result=fn(*args, **kwargs))

        def executor_factory(max_workers):
            executor = FakeProcessPoolExecutor(max_workers, future_factory)
            executors.append(executor)
            return executor

        def fake_as_completed(futures):
            futures = list(futures)
            return [futures[1], futures[0]]

        with patch("darkdump_collector.ProcessPoolExecutor", side_effect=executor_factory), patch(
            "darkdump_collector.as_completed", side_effect=fake_as_completed
        ), patch("sys.stdout", new_callable=io.StringIO) as fake_stdout:
            result = module.batch_collect_dark_net([" alpha ", "beta"], 2)

        self.assertEqual(executors[0].max_workers, 2)
        self.assertEqual(result["requested_amount"], 2)
        self.assertEqual(result["keywords"], ["alpha", "beta"])
        self.assertEqual(result["success_count"], 2)
        self.assertEqual(result["failure_count"], 0)
        self.assertEqual(len(result["items"]), 2)
        self.assertEqual(result["items"][0]["status"], "success")
        self.assertEqual(result["items"][0]["search_keyword"], "alpha")
        self.assertEqual(result["items"][1]["status"], "success")
        self.assertEqual(result["items"][1]["search_keyword"], "beta")

        stdout = fake_stdout.getvalue()
        self.assertLess(stdout.index("query=beta"), stdout.index("query=alpha"))

    def test_batch_collect_dark_net_rejects_invalid_keywords(self):
        module = get_darkdump_collector_module()

        with self.assertRaises(TypeError):
            module.batch_collect_dark_net("alpha", 1)

        with self.assertRaises(ValueError):
            module.batch_collect_dark_net([], 1)

        with self.assertRaises(TypeError):
            module.batch_collect_dark_net(["alpha", 1], 1)

        with self.assertRaises(ValueError):
            module.batch_collect_dark_net(["alpha", "   "], 1)

    @patch("darkdump_collector._batch_collect_worker")
    def test_batch_collect_dark_net_records_errors_and_continues(self, mock_worker):
        module = get_darkdump_collector_module()
        executors = []

        def worker_side_effect(index, key_word, amount, retry_times):
            if key_word == "beta":
                return {
                    "index": index,
                    "item": {
                        "collected_date": "2026-04-08",
                        "collected_time": "10:00:01",
                        "search_keyword": key_word,
                        "status": "error",
                        "error": "tor down",
                    },
                }

            return {
                "index": index,
                "item": {
                    "collected_date": "2026-04-08",
                    "collected_time": f"10:00:0{index}",
                    "search_keyword": key_word,
                    "status": "success",
                    "collect_result": {
                        "query": key_word,
                        "requested_amount": amount,
                        "returned_count": 1,
                        "proxy_enabled": True,
                        "scrape_enabled": True,
                        "images_enabled": False,
                        "tor_checked": True,
                        "tor_ok": True,
                        "tor_ip": f"Current IP Address via Tor: 185.220.101.{index + 1}",
                        "errors": [],
                        "results": [{"index": 1, "title": key_word.title()}],
                    },
                },
            }

        mock_worker.side_effect = worker_side_effect

        def future_factory(fn, *args, **kwargs):
            return FakeFuture(result=fn(*args, **kwargs))

        def executor_factory(max_workers):
            executor = FakeProcessPoolExecutor(max_workers, future_factory)
            executors.append(executor)
            return executor

        def fake_as_completed(futures):
            return list(futures)

        with patch("darkdump_collector.ProcessPoolExecutor", side_effect=executor_factory), patch(
            "darkdump_collector.as_completed", side_effect=fake_as_completed
        ):
            result = module.batch_collect_dark_net(["alpha", "beta", "gamma"], 2)

        self.assertEqual(result["success_count"], 2)
        self.assertEqual(result["failure_count"], 1)
        self.assertEqual(result["items"][1]["status"], "error")
        self.assertEqual(result["items"][1]["search_keyword"], "beta")
        self.assertEqual(result["items"][1]["error"], "tor down")
        self.assertEqual(result["items"][2]["status"], "success")
        self.assertEqual(executors[0].max_workers, 3)

    def test_save_batch_collect_dark_net_to_excel_writes_expected_rows(self):
        from openpyxl import load_workbook

        module = get_darkdump_collector_module()
        batch_result = {
            "requested_amount": 2,
            "keywords": ["alpha", "beta"],
            "success_count": 2,
            "failure_count": 0,
            "items": [
                {
                    "collected_date": "2026-04-08",
                    "collected_time": "10:00:00",
                    "search_keyword": "alpha",
                    "status": "success",
                    "collect_result": {
                        "query": "alpha",
                        "requested_amount": 2,
                        "returned_count": 2,
                        "proxy_enabled": True,
                        "scrape_enabled": True,
                        "images_enabled": False,
                        "tor_checked": True,
                        "tor_ok": True,
                        "tor_ip": "Current IP Address via Tor: 185.220.101.1",
                        "errors": [{"stage": "site", "url": "http://dead.onion", "message": "dead onion"}],
                        "results": [
                            {
                                "index": 1,
                                "title": "Alpha 1",
                                "description": "Primary listing",
                                "onion_link": "http://alpha-1.onion",
                                "keywords": ["alpha", "market"],
                                "sentiment": {"polarity": 0.4, "subjectivity": 0.2},
                                "metadata": {"description": "meta-1"},
                                "links": ["https://example.com"],
                                "link_count": 1,
                                "emails": ["ops@alpha.onion"],
                                "documents": ["docs/report.pdf"],
                            },
                            {
                                "index": 2,
                                "title": "Alpha 2",
                                "description": "Secondary listing",
                                "onion_link": "http://alpha-2.onion",
                                "keywords": ["alpha", "backup"],
                                "sentiment": {"polarity": 0.1, "subjectivity": 0.3},
                                "metadata": {"description": "meta-2"},
                                "links": ["https://example.org"],
                                "link_count": 1,
                                "emails": ["admin@alpha.onion"],
                                "documents": [],
                            },
                        ],
                    },
                },
                {
                    "collected_date": "2026-04-08",
                    "collected_time": "10:00:01",
                    "search_keyword": "beta",
                    "status": "success",
                    "collect_result": {
                        "query": "beta",
                        "requested_amount": 2,
                        "returned_count": 0,
                        "proxy_enabled": True,
                        "scrape_enabled": True,
                        "images_enabled": False,
                        "tor_checked": True,
                        "tor_ok": True,
                        "tor_ip": "Current IP Address via Tor: 185.220.101.2",
                        "errors": [],
                        "results": [],
                    },
                },
            ],
        }

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_file:
            output_path = tmp_file.name

        try:
            saved_path = module.save_batch_collect_dark_net_to_excel(batch_result, output_path)
            workbook = load_workbook(saved_path)
            worksheet = workbook["results"]
            rows = list(worksheet.iter_rows(values_only=True))
        finally:
            if os.path.exists(output_path):
                os.remove(output_path)

        expected_headers = [
            "采集日期",
            "采集时间",
            "搜索关键词",
            "query",
            "requested_amount",
            "returned_count",
            "proxy_enabled",
            "scrape_enabled",
            "images_enabled",
            "tor_checked",
            "tor_ok",
            "tor_ip",
            "errors",
            "index",
            "title",
            "description",
            "onion_link",
            "keywords",
            "sentiment",
            "metadata",
            "links",
            "link_count",
            "emails",
            "documents",
        ]

        self.assertEqual(list(rows[0]), expected_headers)
        self.assertNotIn("results", rows[0])
        self.assertEqual(len(rows), 3)
        self.assertEqual(rows[1][2], "alpha")
        self.assertEqual(rows[1][3], "alpha")
        self.assertEqual(rows[1][11], "Current IP Address via Tor: 185.220.101.1")
        self.assertEqual(rows[1][12], json.dumps([{"stage": "site", "url": "http://dead.onion", "message": "dead onion"}], ensure_ascii=False))
        self.assertEqual(rows[1][14], "Alpha 1")
        self.assertEqual(rows[1][17], json.dumps(["alpha", "market"], ensure_ascii=False))
        self.assertEqual(rows[1][18], json.dumps({"polarity": 0.4, "subjectivity": 0.2}, ensure_ascii=False))
        self.assertEqual(rows[2][14], "Alpha 2")
        self.assertEqual(rows[2][23], json.dumps([], ensure_ascii=False))

    def test_save_batch_collect_dark_net_to_excel_creates_header_only_workbook_for_empty_results(self):
        from openpyxl import load_workbook

        module = get_darkdump_collector_module()
        batch_result = {
            "requested_amount": 2,
            "keywords": ["alpha", "beta"],
            "success_count": 1,
            "failure_count": 1,
            "items": [
                {
                    "collected_date": "2026-04-08",
                    "collected_time": "10:00:00",
                    "search_keyword": "alpha",
                    "status": "success",
                    "collect_result": {
                        "query": "alpha",
                        "requested_amount": 2,
                        "returned_count": 0,
                        "proxy_enabled": True,
                        "scrape_enabled": True,
                        "images_enabled": False,
                        "tor_checked": True,
                        "tor_ok": True,
                        "tor_ip": "Current IP Address via Tor: 185.220.101.1",
                        "errors": [],
                        "results": [],
                    },
                },
                {
                    "collected_date": "2026-04-08",
                    "collected_time": "10:00:01",
                    "search_keyword": "beta",
                    "status": "error",
                    "error": "tor down",
                },
            ],
        }

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_file:
            output_path = tmp_file.name

        try:
            saved_path = module.save_batch_collect_dark_net_to_excel(batch_result, output_path)
            workbook = load_workbook(saved_path)
            worksheet = workbook["results"]
            rows = list(worksheet.iter_rows(values_only=True))
        finally:
            if os.path.exists(output_path):
                os.remove(output_path)

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0][0], "采集日期")

    def test_default_key_words_are_non_empty_strings(self):
        module = get_darkdump_collector_module()

        self.assertEqual(len(module.DEFAULT_KEY_WORDS), len(module.DEFAULT_KEY_WORDS_STR.split(",")))
        self.assertGreater(len(module.DEFAULT_KEY_WORDS), 0)
        self.assertTrue(all(isinstance(key_word, str) and key_word.strip() for key_word in module.DEFAULT_KEY_WORDS))
        self.assertEqual(module.DEFAULT_AMOUNT, 20)

    @patch("darkdump_collector.save_batch_collect_dark_net_to_excel", return_value="/tmp/darkdump_batch_results_20260408_153000.xlsx")
    @patch("darkdump_collector.batch_collect_dark_net")
    def test_main_uses_defaults_and_prints_summary(self, mock_batch_collect, mock_save_excel):
        module = get_darkdump_collector_module()
        mock_batch_collect.return_value = {
            "requested_amount": 20,
            "keywords": module.DEFAULT_KEY_WORDS,
            "success_count": 8,
            "failure_count": 2,
            "items": [],
        }

        with patch("sys.stdout", new_callable=io.StringIO) as fake_stdout:
            result = module.main()

        mock_batch_collect.assert_called_once_with(module.DEFAULT_KEY_WORDS, 20)
        save_args = mock_save_excel.call_args[0]
        self.assertEqual(save_args[0], mock_batch_collect.return_value)
        self.assertRegex(save_args[1], r"^darkdump_batch_results_\d{8}_\d{6}\.xlsx$")

        self.assertEqual(result["key_words"], module.DEFAULT_KEY_WORDS)
        self.assertEqual(result["amount"], 20)
        self.assertEqual(result["batch_result"], mock_batch_collect.return_value)
        self.assertEqual(result["excel_path"], "/tmp/darkdump_batch_results_20260408_153000.xlsx")

        stdout = fake_stdout.getvalue()
        self.assertIn(f"Keyword count: {len(module.DEFAULT_KEY_WORDS)}", stdout)
        self.assertIn("Amount per keyword: 20", stdout)
        self.assertIn("Success count: 8, Failure count: 2", stdout)
        self.assertIn("Excel saved to: /tmp/darkdump_batch_results_20260408_153000.xlsx", stdout)


if __name__ == "__main__":
    unittest.main()
